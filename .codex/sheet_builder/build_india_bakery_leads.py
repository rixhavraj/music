from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Paragraph, Spacer, Table as PdfTable, TableStyle

ROOT = Path("D:/WorkSpace/FullStack/music")
XLSX = ROOT / "outputs/bakery_leads/india_bakery_website_sales_leads.xlsx"
PDF = ROOT / "output/pdf/india_bakery_website_sales_leads.pdf"
XLSX.parent.mkdir(parents=True, exist_ok=True)
PDF.parent.mkdir(parents=True, exist_ok=True)

leads = [
    ["PURI Bakers", "Bakery / cake shop", "011 2808 6566; 088260 54079; 098115 96000", "", "Krishna Plaza, Dwarka Sector 10", "Delhi", "Delhi", "110075", "https://wap.justdial.com/Delhi/Puri-Bakers-Infront-Niit-Near-Central-Market-Dwarka-Sector-10/011PXX11-XX11-130508094519-F9S1_BZDET", "Bakery, cake shop, manufacturers; delivery and in-store collect", "Directory only / no clear site", "High"],
    ["Mr. Jain Bakery & More", "Cake manufacturer", "08123557687", "", "Ghaziabad", "Ghaziabad", "Uttar Pradesh", "", "https://www.justdial.com/Delhi/Cake-Manufacturers/nct-10070023", "Cake manufacturer with online-order listing", "Directory only / no clear site", "High"],
    ["Indian Bakers", "Cake manufacturer", "09008531065", "", "Noida", "Noida", "Uttar Pradesh", "", "https://www.justdial.com/Delhi/Cake-Manufacturers/nct-10070023", "24 years in business; manufacturer listing", "Directory only / no clear site", "High"],
    ["Cake Connoisseur", "Cake manufacturer", "08147113835", "", "Delhi", "Delhi", "Delhi", "", "https://www.justdial.com/Delhi/Cake-Manufacturers/nct-10070023", "Cake manufacturer; event/custom order fit", "Directory only / no clear site", "High"],
    ["Flower Cake Delhi", "Cake manufacturer", "07411719195", "", "Rajouri Garden", "Delhi", "Delhi", "", "https://www.justdial.com/Delhi/Cake-Manufacturers/nct-10070023", "Top pick; high call pickup; custom cakes", "Directory only / no clear site", "High"],
    ["Mercons Foods", "Cake manufacturer", "08123425913", "", "West Punjabi Bagh", "Delhi", "Delhi", "", "https://www.justdial.com/Delhi/Cake-Manufacturers/nct-10070023", "Top pick manufacturer; possible B2B/catering lead", "Directory only / no clear site", "High"],
    ["Masters Cakes and Sweets", "Bakery / sweets", "07947155804", "", "Ashok Vihar 1", "Delhi", "Delhi", "", "https://www.justdial.com/Delhi/Bakeries-in-New-Delhi/nct-10033880", "High-volume bakery listing; sweets and cakes", "Directory only / no clear site", "High"],
    ["The Fairy Bake", "Bakery", "08123821686", "", "Lodhi Road", "Delhi", "Delhi", "", "https://www.justdial.com/Delhi/Bakeries-in-New-Delhi/nct-10033880", "Top pick with high call pickup", "Directory only / no clear site", "High"],
    ["Blue Ribbon", "Bakery", "07383290589", "", "Lajpat Nagar 4", "Delhi", "Delhi", "", "https://www.justdial.com/Delhi/Bakeries-in-New-Delhi/nct-10033880", "Top pick bakery listing", "Directory only / no clear site", "High"],
    ["Appetite German Bakery", "Bakery", "", "", "Main Bazar Road, Chuna Mandi, Pahar Ganj", "Delhi", "Delhi", "110055", "https://www.justdial.com/Delhi/Bakeries-in-New-Delhi/nct-10033880", "Local bakery listing; likely tourist/footfall area", "Directory only / no clear site", "High"],
    ["St Anthonys Bakery & Cafe", "Bakery cafe", "09036695475", "", "Kolbad Road, Thane West", "Thane", "Maharashtra", "400601", "https://www.justdial.com/Mumbai/Bakeries/nct-10033880", "Large rating count; online ordering/listing presence", "Website needs review", "High"],
    ["Baker's Cutlery", "Bakery supplies / bakery", "09972332390", "", "Outside Crawford Market", "Mumbai", "Maharashtra", "400001", "https://www.justdial.com/Mumbai/Bakeries/nct-10033880", "Crawford Market listing; WhatsApp visible", "Directory only / no clear site", "High"],
    ["Edwards The Bakery", "Bakery", "09036288630", "", "Bazargate Street, Fort", "Mumbai", "Maharashtra", "400001", "https://www.justdial.com/Mumbai/Bakeries/nct-10033880", "Heritage/fort area bakery; online-order listing", "Website needs review", "High"],
    ["Bangalore Iyengar Bakery", "Bakery", "09036276020", "", "MG Road, Ghatkopar East", "Mumbai", "Maharashtra", "400077", "https://www.justdial.com/Mumbai/Bakeries/nct-10033880", "Vegetarian bakery; local search opportunity", "Directory only / no clear site", "High"],
    ["KKdict", "Bakery", "09725304505", "", "Pokhran Road, Upvan", "Thane", "Maharashtra", "400606", "https://www.justdial.com/Mumbai/Bakeries/nct-10033880", "Small bakery listing with visible phone", "Directory only / no clear site", "High"],
    ["Nil's Nature's Bake Pvt Ltd", "Bakery", "09972204007", "", "Andheri Kurla Road, Sakinaka", "Mumbai", "Maharashtra", "400072", "https://www.justdial.com/Mumbai/Bakeries/nct-10033880", "High inquiry count; WhatsApp listing", "Website needs review", "High"],
    ["Meher Sans Food Plaza", "Bakery / food plaza", "08401123596", "", "Mangel Ali Murbe", "Palghar", "Maharashtra", "401501", "https://www.justdial.com/Mumbai/Bakeries/nct-10033880", "GST verified listing; fast-response box", "Directory only / no clear site", "High"],
    ["Bakery House", "Bakery / cake shop", "", "", "Shop No 19, Market Road, Andheri West", "Mumbai", "Maharashtra", "400061", "https://www.justdial.com/Mumbai/Bakery-House-Opposite-Mahalaxmi-Jwellers-Andheri-West/022PXX22-XX22-140205204307-X3X7_BZDET", "38 years in business; custom order services", "Directory only / no clear site", "High"],
    ["Mohammadi Bakery", "Bakery product retailer", "", "", "Near Police Chowki, Chandni Agar, S P Road, Wadala East", "Mumbai", "Maharashtra", "400037", "https://www.justdial.com/Mumbai/Mohammadi-Bakery-Near-Police-Chowki-Chandni-Agar-Wadala-East/022PXX22-XX22-230207155019-L8K8_BZDET", "Small local bakery listing", "Directory only / no clear site", "High"],
    ["Bombay Bakery", "Bakery equipment / bakery", "", "", "Gulshan Market, Khairani Road, Sakinaka", "Mumbai", "Maharashtra", "400072", "https://www.justdial.com/jdmart/Mumbai/Bombay-Bakery-Sakinaka/022PXX22-XX22-180626151445-P1K3_BZDET/catalogue", "JD Mart listing has Add Website prompt", "Directory only / no clear site", "High"],
    ["Kalyani Bakery And Confectionery", "Bakery / confectionery", "", "", "381/3 Prabhune Road, Narayan Peth", "Pune", "Maharashtra", "411030", "https://www.justdial.com/Pune/Kalyani-Bakery-And-Confectionery-Opposite-Kanya-Shala-Narayan-Peth/020PXX20-XX20-170208063828-Q9X1_BZDET", "Small listing; custom-order opportunity", "Directory only / no clear site", "High"],
    ["Swiss Castle Bakery", "Bakery / cake supplier", "", "", "Golconda X Road, Musheerabad", "Hyderabad", "Telangana", "500020", "https://m.justdial.com/jdmart/Hyderabad/Cake/jdm-1013374-ent-2-6848568", "Cake supplier listing on JD Mart", "Directory only / no clear site", "High"],
    ["Shivanand Bakers", "Bakery / cake supplier", "", "", "Indira Nagar, Dilsukh Nagar", "Hyderabad", "Telangana", "500060", "https://m.justdial.com/jdmart/Hyderabad/Cake/jdm-1013374-ent-2-6848568", "Cake supplier listing; high rating in snippet", "Directory only / no clear site", "High"],
    ["Nainas Pastries And Bake", "Bakery", "099645 95975", "", "Bengaluru", "Bengaluru", "Karnataka", "", "https://nainaspastries-and-bake.ueniweb.com/", "UENI-hosted site; likely template site", "Website needs review", "Medium"],
    ["Cakewala", "Bakery / cakes", "080-42066037; 8884420929", "cakewala2@gmail.com", "No. 54/A, 24th Main, 2nd Phase, J P Nagar", "Bengaluru", "Karnataka", "", "https://www.storeboard.com/cakewala2", "Directory profile references a website; verify current site quality", "Website needs review", "High"],
    ["Amma's Pastries", "Bakery", "(080) 25424330; 09945192403; 09590607750", "ammaspastries@yahoo.co.in", "1, CM Complex, 80 Feet Road, Banaswadi", "Bengaluru", "Karnataka", "560033", "https://www.sg.jupiteryellowdetail.com/Bangalore/search-by-listings/bakery-/banaswadi-/ammaand39s-pastries/314078.jws", "Directory profile with contact details", "Website needs review", "Medium"],
    ["Bengaluru Bread Club", "Bakery manufacturer", "+91-9611288610; +91-8970713715", "info@bengalurubreadclub.com", "BM Shankrappa Industrial Estate, Sunkadakatte", "Bengaluru", "Karnataka", "560091", "https://bengalurubreadclub.in/contact", "Manufacturing bakery products; custom orders", "Working website", "Medium"],
]


def pitch(row):
    name, typ, *_rest, angle, status, priority = row
    city = row[5]
    if status == "Directory only / no clear site":
        return f"Pitch {name} a first proper website: menu/product gallery, WhatsApp/call buttons, Google Maps, inquiry form, reviews, and local SEO for {city}. They likely want more direct calls and less dependence on Justdial/listing pages."
    if status == "Website needs review":
        return f"Pitch a quick audit for {name}: check if ordering, mobile speed, WhatsApp buttons, product photos, and Google search visibility are costing them orders. They likely want more online cake/custom-order inquiries."
    if "manufacturer" in typ.lower() or "supplier" in typ.lower():
        return f"Pitch B2B lead capture: wholesale catalog, quote form, service area page, and repeat-order workflow. They likely want restaurant, cafe, and event buyers."
    return f"Pitch a conversion refresh: mobile menu, photos, WhatsApp ordering, event/catering pages, reviews, and local SEO. They likely want more orders from nearby customers."


headers = ["Business Name", "Business Type", "Phone", "Email", "Address", "City", "State", "ZIP", "Website / Source", "Sales Angle", "Website Status", "Priority", "Outreach Status", "How To Pitch / What They Want", "Notes"]

wb = Workbook()
ws = wb.active
ws.title = "India Bakery Leads"
summary = wb.create_sheet("Summary")
ws.append(headers)
for row in leads:
    ws.append(row[:12] + ["New", pitch(row), ""])

header_fill = PatternFill("solid", fgColor="234E52")
alt_fill = PatternFill("solid", fgColor="EAF7FB")
thin = Side(style="thin", color="CBD5E1")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(bottom=thin)
        if cell.row % 2 == 0:
            cell.fill = alt_fill
widths = [28, 22, 24, 24, 34, 16, 16, 10, 54, 42, 22, 12, 16, 62, 30]
for idx, width in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + idx)].width = width
ws.freeze_panes = "A2"
ws.add_table(Table(displayName="IndiaBakeryLeadsTable", ref=f"A1:O{len(leads)+1}", tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
ws.add_data_validation(DataValidation(type="list", formula1='"High,Medium,Low"', sqref=f"L2:L{len(leads)+1}"))
ws.add_data_validation(DataValidation(type="list", formula1='"New,Called,Emailed,Follow-up,Won,Not Fit"', sqref=f"M2:M{len(leads)+1}"))

summary["A1"] = "India Bakery Website Sales Leads"
summary["A1"].font = Font(bold=True, size=18, color="234E52")
summary["A2"] = "Focus: India bakeries with no clear website, directory-only presence, or website-review opportunities."
summary["A4"], summary["B4"] = "Metric", "Value"
metrics = [
    ("Total leads", f"=COUNTA('India Bakery Leads'!A2:A{len(leads)+1})"),
    ("High priority leads", f'=COUNTIF(\'India Bakery Leads\'!L2:L{len(leads)+1},"High")'),
    ("No clear website / directory only", f'=COUNTIF(\'India Bakery Leads\'!K2:K{len(leads)+1},"Directory only / no clear site")'),
    ("Website needs review", f'=COUNTIF(\'India Bakery Leads\'!K2:K{len(leads)+1},"Website needs review")'),
    ("Visible phone numbers", f"=COUNTA('India Bakery Leads'!C2:C{len(leads)+1})"),
]
for r, (label, formula) in enumerate(metrics, 5):
    summary[f"A{r}"], summary[f"B{r}"] = label, formula
for row in summary.iter_rows(min_row=4, max_row=9, max_col=2):
    for cell in row:
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(wrap_text=True)
for cell in summary[4]:
    cell.fill = header_fill
    cell.font = Font(color="FFFFFF", bold=True)
summary.column_dimensions["A"].width = 32
summary.column_dimensions["B"].width = 16
wb.save(XLSX)

styles = getSampleStyleSheet()
title = ParagraphStyle("Title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, textColor=colors.HexColor("#234E52"))
body = ParagraphStyle("Body", parent=styles["BodyText"], fontSize=7, leading=9)
small = ParagraphStyle("Small", parent=body, fontSize=6.3, leading=8)
hdr = ParagraphStyle("Hdr", parent=body, fontName="Helvetica-Bold", textColor=colors.white)

def para(text, style=body):
    text = str(text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return Paragraph(text, style)

doc = BaseDocTemplate(str(PDF), pagesize=landscape(letter), leftMargin=.35*inch, rightMargin=.35*inch, topMargin=.35*inch, bottomMargin=.45*inch)
frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
doc.addPageTemplates([PageTemplate(id="p", frames=[frame])])
story = [Paragraph("India Bakery Website Sales Leads", title), para("Directory-only, no-clear-website, and website-review bakery prospects across Delhi NCR, Mumbai, Pune, Hyderabad, and Bengaluru."), Spacer(1, .12*inch)]
pdf_headers = ["Business", "Phone", "City", "Website Status", "Priority", "Sales Angle", "How To Pitch / What They Want"]
pdf_data = [[para(h, hdr) for h in pdf_headers]]
for row in leads:
    pdf_data.append([para(row[0], small), para(row[2], small), para(row[5], small), para(row[10], small), para(row[11], small), para(row[9], small), para(pitch(row), small)])
t = PdfTable(pdf_data, repeatRows=1, colWidths=[1.45*inch, 1.05*inch, .75*inch, 1.18*inch, .52*inch, 1.65*inch, 3.23*inch])
t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#234E52")),("GRID",(0,0),(-1,-1),.25,colors.HexColor("#CBD5E1")),("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),3),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#EAF7FB")])]))
story.append(t)
story.append(Spacer(1, .15*inch))
story.append(Paragraph("Source Links", title))
src = [[para("Business", hdr), para("Website / Source", hdr)]] + [[para(r[0], small), para(r[8], small)] for r in leads]
st = PdfTable(src, repeatRows=1, colWidths=[2.6*inch, 7.2*inch])
st.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#234E52")),("GRID",(0,0),(-1,-1),.25,colors.HexColor("#CBD5E1")),("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),3)]))
story.append(st)
doc.build(story)
print(XLSX)
print(PDF)
