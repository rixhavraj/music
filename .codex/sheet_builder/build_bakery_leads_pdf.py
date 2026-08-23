from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)


ROOT = Path("D:/WorkSpace/FullStack/music")
XLSX = ROOT / "outputs/bakery_leads/bakery_website_sales_leads.xlsx"
OUT_DIR = ROOT / "output/pdf"
PDF = OUT_DIR / "bakery_website_sales_leads.pdf"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


OUT_DIR.mkdir(parents=True, exist_ok=True)

wb = load_workbook(XLSX, data_only=True)
lead_ws = wb["Bakery Leads"]
rows = list(lead_ws.iter_rows(values_only=True))
headers = [clean(v) for v in rows[0]]
records = [dict(zip(headers, row)) for row in rows[1:] if row and row[0]]

total = len(records)
high = sum(1 for r in records if clean(r.get("Priority")) == "High")
no_site = sum(1 for r in records if clean(r.get("Website Status")) == "Directory only / no clear site")
review = sum(1 for r in records if clean(r.get("Website Status")) == "Website needs review")
working = sum(1 for r in records if clean(r.get("Website Status")) == "Working website")

styles = getSampleStyleSheet()
title = ParagraphStyle(
    "Title",
    parent=styles["Title"],
    fontName="Helvetica-Bold",
    fontSize=20,
    leading=24,
    textColor=colors.HexColor("#234E52"),
    spaceAfter=10,
)
subtitle = ParagraphStyle(
    "Subtitle",
    parent=styles["BodyText"],
    fontName="Helvetica",
    fontSize=9,
    leading=12,
    textColor=colors.HexColor("#475569"),
    spaceAfter=14,
)
body = ParagraphStyle(
    "Body",
    parent=styles["BodyText"],
    fontName="Helvetica",
    fontSize=7,
    leading=9,
    alignment=TA_LEFT,
)
small = ParagraphStyle(
    "Small",
    parent=body,
    fontSize=6.4,
    leading=8,
)
header_style = ParagraphStyle(
    "Header",
    parent=body,
    fontName="Helvetica-Bold",
    fontSize=6.8,
    leading=8,
    textColor=colors.white,
)


def p(text, style=body):
    text = clean(text)
    text = (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )
    return Paragraph(text, style)


def footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#64748B"))
    canvas.drawString(0.45 * inch, 0.28 * inch, "Bakery Website Sales Leads")
    canvas.drawRightString(10.55 * inch, 0.28 * inch, f"Page {doc.page}")
    canvas.restoreState()


doc = BaseDocTemplate(
    str(PDF),
    pagesize=landscape(letter),
    rightMargin=0.35 * inch,
    leftMargin=0.35 * inch,
    topMargin=0.35 * inch,
    bottomMargin=0.45 * inch,
)
frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
doc.addPageTemplates([PageTemplate(id="lead_pages", frames=[frame], onPage=footer)])

story = []
story.append(Paragraph("Bakery Website Sales Leads", title))
story.append(
    Paragraph(
        "PDF version of the outreach workbook. Focus: bakeries with public phone numbers, "
        "no-clear-website targets, website-review targets, and ordering/store opportunities.",
        subtitle,
    )
)

summary_data = [
    [p("Metric", header_style), p("Value", header_style)],
    [p("Total leads"), p(total)],
    [p("High priority leads"), p(high)],
    [p("No clear website / directory only"), p(no_site)],
    [p("Website needs review"), p(review)],
    [p("Working website"), p(working)],
]
summary_table = Table(summary_data, colWidths=[2.4 * inch, 1.0 * inch], hAlign="LEFT")
summary_table.setStyle(
    TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#234E52")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]
    )
)
story.append(summary_table)
story.append(Spacer(1, 0.16 * inch))
story.append(Paragraph("Lead Table", title))

table_headers = [
    "Business",
    "Phone",
    "City",
    "Website Status",
    "Priority",
    "Sales Angle",
    "How To Pitch / What They Want",
]
table_data = [[p(h, header_style) for h in table_headers]]
for r in records:
    table_data.append(
        [
            p(r.get("Business Name"), small),
            p(r.get("Phone"), small),
            p(r.get("City"), small),
            p(r.get("Website Status"), small),
            p(r.get("Priority"), small),
            p(r.get("Sales Angle"), small),
            p(r.get("How To Pitch / What They Want"), small),
        ]
    )

lead_table = Table(
    table_data,
    repeatRows=1,
    colWidths=[1.42 * inch, 0.9 * inch, 0.72 * inch, 1.15 * inch, 0.55 * inch, 1.65 * inch, 3.42 * inch],
)
lead_table.setStyle(
    TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#234E52")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("PADDING", (0, 0), (-1, -1), 3.2),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF7FB")]),
        ]
    )
)
story.append(lead_table)

story.append(PageBreak())
story.append(Paragraph("Source Links", title))
source_data = [[p("Business", header_style), p("Website / Source", header_style)]]
for r in records:
    source_data.append([p(r.get("Business Name"), small), p(r.get("Website / Source"), small)])
source_table = Table(source_data, repeatRows=1, colWidths=[2.8 * inch, 7.0 * inch])
source_table.setStyle(
    TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#234E52")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("PADDING", (0, 0), (-1, -1), 3.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]
    )
)
story.append(source_table)

doc.build(story)
print(PDF)
